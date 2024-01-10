# -*- coding: utf-8 -*-            
# @Author : ben
# @Time : 2023/12/13 11:30


import openpyxl

class Data:

    def getdata(self):

        # 获取工作簿
        wb = openpyxl.load_workbook('Antman.xlsx')

        # 获取工作表
        sheet1 = wb['Antman']

        # 获取工作单元格
        # print(sheet1.cell(row=11, column=3).value)
        #
        # # 获取最大行数、列数
        # print(sheet1.max_row)
        #
        # print(sheet1.max_column)

        rows = sheet1.max_row

        cols = sheet1.max_column

        datalist = []

        for x in range(1, rows + 1):
            dict1 = {}
            for y in range(1, cols + 1):
                # print(sheet1.cell(x,y).value)
                values = sheet1.cell(x, y).value
                dict1[sheet1.cell(1, y).value] = values
                datalist.append(dict1)
        print(datalist)

        return datalist



